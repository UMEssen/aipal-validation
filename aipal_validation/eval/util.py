import pandas as pd
import yaml
import os
import numpy as np
# Add openpyxl for Excel formatting
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
from sklearn.metrics import roc_curve

# Get the project root directory (two levels up from this file)
PROJECT_ROOT = os.path.dirname(os.path.dirname(os.path.dirname(os.path.abspath(__file__))))

def load_data(config_path='/home/merengelke/aipal_validation/aipal_validation/config/config_outlier.yaml', root_path='/local/work/merengelke/aipal/', filter_by_size=True, is_adult=True, filter_missing_values=True):
    """
    Load and preprocess data from multiple cohorts based on configuration.

    Parameters:
    -----------
    config_path : str
        Path to the configuration YAML file
    root_path : str
        Root path for data files
    filter_by_size : bool
        Whether to filter cohorts by minimum size (default: True)
    is_adult : bool
        Whether to load adult (True) or children (False) data. This controls both
        the age filtering and which cities to include in the analysis.

    Returns:
    --------
    pandas.DataFrame
        Preprocessed dataframe with all cohorts
    dict
        Configuration dictionary
    list
        List of feature columns
    """
    # Load configuration
    with open(config_path, 'r') as f:
        config = yaml.load(f, Loader=yaml.FullLoader)

    config['is_adult'] = is_adult

    # Get paths for all cohorts
    cities_countries = []
    # Try different possible keys that might exist in the config
    for key in ['cities_countries', 'cities_country', 'city_countries', 'city_country']:
        if key in config:
            cities_countries = config[key]
            break

    # If no cities found, use default
    if not cities_countries:
        print("Warning: No cities/countries found in config file. Using default list.")
        cities_countries = ['dallas', 'melbourne', 'maastricht', 'barcelona', 'bochum',
                           'wroclaw', 'kolkata', 'rome', 'salamanca', 'sao_paulo', 'turkey',
                           'buenos_aires', 'milano', 'suzhou', 'antananarivo', 'hannover',
                           'newcastle', 'lagos', 'Vessen']

    # Only filter out certain cities for adult data
    if is_adult:
        cities_countries = [city_country for city_country in cities_countries if city_country != 'newcastle' and city_country != 'turkey' and city_country != 'Vessen']

    paths = [f"{root_path.rstrip('/')}/{city_country}/aipal/predict.csv" for city_country in cities_countries]

    # Load and concatenate data
    df = pd.DataFrame()
    for path in paths:
        try:
            df_small = pd.read_csv(path)
            df_small['city_country'] = path.split('/')[-3]
            df = pd.concat([df, df_small])
        except Exception as e:
            print(f"Error loading {path}: {e}")

    # Filter by age based on is_adult parameter
    if is_adult:
        print("Filtering adults")
        df = df[df['age'] > 18]
    else:
        print("Filtering children")
        df = df[df['age'] <= 18]

    # remove samples with more than 20 % missing values
    if filter_missing_values:
        # Check for feature_columns key
        feature_columns = []
        for key in ['feature_columns', 'features', 'input_features']:
            if key in config:
                feature_columns = config[key]
                break

        # If no feature columns found, use default list
        if not feature_columns:
            print("Warning: No feature columns found in config file. Using default list.")
            feature_columns = ['Fibrinogen_g_L', 'MCV_fL', 'Monocytes_percent', 'LDH_UI_L',
                                'PT_percent', 'MCHC_g_L', 'Lymphocytes_G_L', 'age',
                                'Monocytes_G_L', 'Platelets_G_L']

        # Store the feature columns in config for later use
        config['feature_columns'] = feature_columns

        # Filter out samples with more than 20 % missing values
        df_input_features = df[feature_columns]
        missing_samples = df_input_features.isna().sum(axis=1) > 0.2 * len(feature_columns)
        df = df[~missing_samples]


    df['city_country'] = df['city_country'].str.replace('_', ' ')
    df['city_country'] = df['city_country'].str.capitalize()

    # map M and F to Male and Female
    df['sex'] = df['sex'].replace({'M': 'Male', 'F': 'Female'})
    df['sex'] = df['sex'].replace('I', 'Male')

    # Drop unnecessary columns
    df.drop(columns=['ELN', 'Diagnosis', 'additional.diagnosis.details..lineage.etc', 'lineage.details', 'birth_date', 'ID',
                     'encounter_start', 'encounter_end', 'patient_id', 'condition_codes', 'condition_id', 'encounter_id',
                     'recorded_date'
                     ],
            inplace=True, errors='ignore')

    # Clean class values
    df['class'] = df['class'].str.strip()

    # Filter cohorts by size if requested
    if filter_by_size:
        df = df.groupby('city_country').filter(lambda x: len(x) > 30)

    # Get feature columns from config
    features = config['feature_columns']  # This is safe now because we ensured it exists above

    return df, config, features

def save_to_excel(df_table, output_path):
    """Save the analysis table to a formatted Excel file with features as rows and leukemia types as columns."""
    # Create a Pandas Excel writer using openpyxl
    writer = pd.ExcelWriter(output_path, engine='openpyxl')

    # Convert the dataframe to an Excel object
    df_table.to_excel(writer, sheet_name='Analysis', index=False)

    # Get the worksheet
    worksheet = writer.sheets['Analysis']

    # Define styles
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
    feature_font = Font(bold=True)
    feature_fill = PatternFill(start_color="DBE5F1", end_color="DBE5F1", fill_type="solid")
    reference_font = Font(italic=True, color="666666")
    reference_fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
    combined_font = Font(bold=True)
    type_font = Font(italic=True)
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    # Format headers with a bold blue style and ensure proper alignment
    for col_num, column_title in enumerate(df_table.columns, 1):
        cell = worksheet.cell(row=1, column=col_num)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        cell.border = border

    # Format data cells and adjust column widths
    column_widths = {}

    # First pass to calculate optimal column widths
    for col_idx, column in enumerate(df_table.columns, 1):
        column_letter = get_column_letter(col_idx)
        # For disease type columns (may include n values), set a fixed width
        if any(class_type in column for class_type in ['ALL', 'AML', 'APL']):
            column_widths[column_letter] = 20
        elif column == 'Variable':
            column_widths[column_letter] = 18
        elif column == 'Type':
            column_widths[column_letter] = 15
        else:
            # For other columns, calculate based on content
            max_length = max(
                len(str(column)),
                df_table[column].astype(str).map(len).max()
            )
            column_widths[column_letter] = min(max_length + 2, 25)

    # Apply calculated column widths
    for col_letter, width in column_widths.items():
        worksheet.column_dimensions[col_letter].width = width

    # Format rows
    prev_variable = None
    for row_idx in range(2, worksheet.max_row + 1):
        variable = worksheet.cell(row=row_idx, column=1).value
        type_cell = worksheet.cell(row=row_idx, column=2)

        # Style format for feature rows (with variable name)
        if variable and variable.strip() and variable != prev_variable:
            # This is a feature row - bold it and color the background
            for col_idx in range(1, worksheet.max_column + 1):
                cell = worksheet.cell(row=row_idx, column=col_idx)
                cell.font = feature_font
                cell.fill = feature_fill
            prev_variable = variable

        # Style reference vs combined values differently
        if type_cell.value:
            type_value = type_cell.value.strip()
            if type_value.startswith("Reference"):
                # Format reference rows with gray italic text
                for col_idx in range(1, worksheet.max_column + 1):
                    cell = worksheet.cell(row=row_idx, column=col_idx)
                    cell.font = reference_font
                    cell.fill = reference_fill
            elif type_value == "Mean (sd)" or type_value == "Median [IQR]":
                # Format combined data rows with bold text
                for col_idx in range(1, worksheet.max_column + 1):
                    cell = worksheet.cell(row=row_idx, column=col_idx)
                    if col_idx >= 3:  # Only the actual data cells, not the labels
                        cell.font = combined_font
                    else:
                        cell.font = type_font
            else:
                # Make the type cell italic
                type_cell.font = type_font

        # Center-align all cells and add borders
        for col_idx in range(1, worksheet.max_column + 1):
            cell = worksheet.cell(row=row_idx, column=col_idx)
            # Align differently based on column
            if col_idx == 1:  # Variable column
                cell.alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
            elif col_idx == 2:  # Type column
                cell.alignment = Alignment(horizontal='right', vertical='center', wrap_text=True)
            else:  # Data columns (ALL, AML, APL)
                cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            cell.border = border

    # Add thick horizontal lines between feature groups
    prev_variable = None
    for row_idx in range(2, worksheet.max_row + 1):
        variable = worksheet.cell(row=row_idx, column=1).value

        # Add a thicker border above feature rows
        if variable and variable.strip() and variable != prev_variable and row_idx > 2:
            for col_idx in range(1, worksheet.max_column + 1):
                cell = worksheet.cell(row=row_idx, column=col_idx)
                thick_border = Border(
                    left=Side(style='thin'),
                    right=Side(style='thin'),
                    top=Side(style='thick'),
                    bottom=Side(style='thin')
                )
                cell.border = thick_border

        prev_variable = variable if variable else prev_variable

    # Freeze the header row and first two columns
    worksheet.freeze_panes = 'C2'

    # Save the workbook
    writer.close()

def post_filter(data, logger=None):
    """
    Apply clinical filters to flag samples based on neutrophil and WBC criteria.
    These filters help identify cases that are less likely or more likely to be leukemia.

    Parameters:
    -----------
    data : dict or pandas.DataFrame
        Single sample as dict or DataFrame with multiple samples
    logger : logging.Logger, optional
        Logger for output. If None, no logging is performed.

    Returns:
    --------
    dict or pandas.DataFrame
        Input data with added post_filter_flag and post_filter_outlier columns
    """
    # Handle single sample (dict input)
    if isinstance(data, dict):
        result_df = pd.DataFrame([data])
        is_single_sample = True
    else:
        result_df = data.copy()
        is_single_sample = False

    if logger:
        logger.info("Applying post-filters based on clinical criteria...")

    # Initialize filter flags column
    result_df['post_filter_flag'] = ''
    result_df['post_filter_outlier'] = 0

    for idx, row in result_df.iterrows():
        wbc = row.get('WBC_G_L', np.nan)
        neutrophil_abs = row.get('Neutrophils_per_microliter', np.nan)
        neutrophil_pct = row.get('Neutrophil ratio', np.nan)

        # Convert neutrophil ratio to percentage if it's in decimal form
        if pd.notna(neutrophil_pct) and neutrophil_pct <= 1.0:
            neutrophil_pct = neutrophil_pct * 100

        # Calculate neutrophil percentage from absolute if not available
        if pd.isna(neutrophil_pct) and pd.notna(neutrophil_abs) and pd.notna(wbc):
            neutrophil_pct = (neutrophil_abs / (wbc * 1000)) * 100

        # Convert neutrophil_per_microliter to G/L for comparison
        if pd.notna(neutrophil_abs):
            neutrophil_abs_gl = neutrophil_abs / 1000
        else:
            neutrophil_abs_gl = np.nan

        flags = []

        # Filter 1: WBC > 20.0 AND neutrophils % > 20% = less likely leukemia
        if pd.notna(wbc) and pd.notna(neutrophil_pct):
            if wbc > 20.0 and neutrophil_pct > 20:
                flags.append('WBC>20_Neut>20%:less_likely_leukemia')
                result_df.loc[idx, 'post_filter_outlier'] = 1

        # Filter 2: neutrophils > 80% = not acute leukemia, possibly Sepsis or reactive
        if pd.notna(neutrophil_pct):
            if neutrophil_pct > 80:
                flags.append('Neut>80%:not_acute_leukemia_Sepsis')
                result_df.loc[idx, 'post_filter_outlier'] = 1

        # Filter 3: neutrophil absolute count > 20.5 G/L = unlikely leukemia
        if pd.notna(neutrophil_abs_gl):
            if neutrophil_abs_gl > 20.5:
                flags.append('NeutAbs>20.5:unlikely_leukemia')
                result_df.loc[idx, 'post_filter_outlier'] = 1

        # Filter 4: Mild-moderate leukocytosis with high normal platelets
        # Rationale: AL usually has low platelets; high platelets + elevated WBC suggests infection
        # Slightly relaxed platelet threshold to catch a few more cases
        platelets = row.get('Platelets_G_L', np.nan)
        if pd.notna(wbc) and pd.notna(platelets):
            if 12 <= wbc <= 18 and platelets > 150:
                flags.append('WBC_12-18_Plt>150:likely_infection_not_AL')
                result_df.loc[idx, 'post_filter_outlier'] = 1

        # Filter 5: WBC < 4.0 AND neutrophils > 40% = less likely leukemia
        if pd.notna(wbc) and pd.notna(neutrophil_pct):
            if wbc < 3.0 and neutrophil_pct > 50:
                flags.append('WBC<4_Neut>40%:less_likely_leukemia')
                result_df.loc[idx, 'post_filter_outlier'] = 1

        # Filter 6: Neutropenia and higher platelets to keep those with low platelets among the AL suspects.
        # Rationale: Leukopenic AL usually has low platelets and low neutros, this is also true for AA and MDS and malignant marrow Infiltration.
        platelets = row.get('Platelets_G_L', np.nan)
        if pd.notna(neutrophil_abs_gl) and pd.notna(platelets):
            if neutrophil_abs_gl < 1.0 and platelets > 150:
                flags.append('NeutAbs<1.0_Plt>150:unlikely_AL')
                result_df.loc[idx, 'post_filter_outlier'] = 1

        # Combine all flags
        if flags:
            result_df.loc[idx, 'post_filter_flag'] = '; '.join(flags)

    # Log summary if logger provided
    if logger:
        total_flagged = result_df['post_filter_outlier'].sum()
        logger.info(f"Post-filter flagged {total_flagged}/{len(result_df)} samples ({total_flagged/len(result_df)*100:.1f}%)")

        # Log breakdown by filter type
        filter_counts = {}
        for flag_str in result_df['post_filter_flag']:
            if flag_str:
                for flag in flag_str.split('; '):
                    filter_counts[flag] = filter_counts.get(flag, 0) + 1

        logger.info("Breakdown by filter type:")
        for filter_type, count in filter_counts.items():
            logger.info(f"  {filter_type}: {count} samples")

    # Return single sample dict if input was dict
    if is_single_sample:
        return result_df.iloc[0].to_dict()
    else:
        return result_df


def extract_roc_source_data(data, class_name, prediction_col=None):
    """
    Extract source data for ROC curve plotting.

    Parameters:
    -----------
    data : pandas.DataFrame
        DataFrame containing the samples with true labels and predictions
    class_name : str
        The class name to analyze (e.g., 'AML', 'ALL', 'APL')
    prediction_col : str, optional
        Column name for predictions. If None, uses f'prediction.{class_name}'

    Returns:
    --------
    pandas.DataFrame
        DataFrame with columns: sample_id, true_label, predicted_score, TPR, FPR
    """
    if prediction_col is None:
        prediction_col = f'prediction.{class_name}'

    # Check if required columns exist
    if 'class' not in data.columns or prediction_col not in data.columns:
        raise ValueError(f"Required columns 'class' and '{prediction_col}' not found in data")

    # Create binary labels for this class (1 if matches class_name, 0 otherwise)
    y_true = (data['class'] == class_name).astype(int)

    # Get predicted scores
    y_score = data[prediction_col].values

    # Calculate ROC curve to get thresholds, FPR, TPR
    fpr, tpr, thresholds = roc_curve(y_true, y_score)

    # Create a dataframe with sample data
    df_source = pd.DataFrame({
        'sample_id': range(1, len(data) + 1),
        'true_label': y_true,
        'predicted_score': y_score
    })

    # For each sample, find the TPR and FPR at the threshold closest to its predicted score
    # We'll use the thresholds from roc_curve to map each sample to its corresponding TPR/FPR
    df_source['TPR'] = 0.0
    df_source['FPR'] = 0.0

    # Sort by predicted score descending (same as roc_curve does)
    df_sorted = df_source.sort_values('predicted_score', ascending=False).reset_index(drop=True)

    # Calculate cumulative TPR and FPR
    n_positives = y_true.sum()
    n_negatives = len(y_true) - n_positives

    cum_tp = 0
    cum_fp = 0

    for i, row in df_sorted.iterrows():
        if row['true_label'] == 1:
            cum_tp += 1
        else:
            cum_fp += 1

        # TPR = cumulative true positives / total positives
        tpr_val = cum_tp / n_positives if n_positives > 0 else 0
        # FPR = cumulative false positives / total negatives
        fpr_val = cum_fp / n_negatives if n_negatives > 0 else 0

        df_sorted.loc[i, 'TPR'] = tpr_val
        df_sorted.loc[i, 'FPR'] = fpr_val

    # Sort back to original order
    df_source = df_sorted.sort_values('sample_id').reset_index(drop=True)

    return df_source


def save_roc_source_data_to_excel(datasets, plots_dir, tag, classes=["AML", "APL", "ALL"], custom_prediction_cols=None):
    """
    Extract and save ROC source data to Excel files for multiple datasets.

    Parameters:
    -----------
    datasets : dict
        Dictionary mapping dataset keys to (dataset_name, dataframe) tuples
    plots_dir : str
        Directory to save the plots and Excel files
    tag : str
        Tag for the analysis (e.g., 'adult', 'pediatric')
    classes : list
        List of class names to analyze (default: ["AML", "APL", "ALL"])
    custom_prediction_cols : dict, optional
        Dictionary mapping class names to custom prediction column names.
        If None, uses f'prediction.{class_name}' for each class.

    Returns:
    --------
    None
    """
    print("Extracting ROC source data...")
    roc_source_dir = os.path.join(plots_dir, 'roc_source_data')
    os.makedirs(roc_source_dir, exist_ok=True)

    for dataset_key, (dataset_name, df_data) in datasets.items():
        if df_data is None or df_data.empty:
            print(f"Skipping {dataset_name} - no data available")
            continue

        print(f"Processing {dataset_name} ({len(df_data)} samples)")

        # Create Excel file with multiple sheets for each class
        excel_path = os.path.join(roc_source_dir, f'roc_source_{tag.lower()}_{dataset_key}.xlsx')

        with pd.ExcelWriter(excel_path, engine='openpyxl') as writer:
            for class_name in classes:
                try:
                    # Determine prediction column name
                    prediction_col = custom_prediction_cols.get(class_name, f'prediction.{class_name}') if custom_prediction_cols else f'prediction.{class_name}'

                    # Extract source data for this class
                    source_df = extract_roc_source_data(df_data, class_name, prediction_col)
                    # Save to Excel sheet
                    source_df.to_excel(writer, sheet_name=class_name, index=False)
                    print(f"Saved {class_name} data for {dataset_name}: {len(source_df)} samples")
                except Exception as e:
                    print(f"Error extracting ROC source data for {class_name} in {dataset_name}: {e}")

        print(f"Saved ROC source data to: {excel_path}")

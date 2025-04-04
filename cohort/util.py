import pandas as pd
import yaml
import os
# Add openpyxl for Excel formatting
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

# Get the project root directory (two levels up from this file)
PROJECT_ROOT = os.path.dirname(os.path.dirname(os.path.dirname(os.path.abspath(__file__))))

def load_data(config_path='/home/merengelke/aipal_validation/aipal_validation/config/config_analysis.yaml', root_path='/local/work/merengelke/aipal/', filter_by_size=True, is_adult=True):
    """
    Load and preprocess data from multiple cohorts based on configuration.

    Parameters:
    -----------
    config_path : str
        Path to the configuration YAML file
    root_path : str
        Root path for data files
    filter_by_size : bool
        Whether to filter cohorts by minimum size (default: True)
    is_adult : bool
        Whether to load adult (True) or children (False) data. This controls both
        the age filtering and which cities to include in the analysis.

    Returns:
    --------
    pandas.DataFrame
        Preprocessed dataframe with all cohorts
    dict
        Configuration dictionary
    list
        List of feature columns
    """
    # Load configuration
    with open(config_path, 'r') as f:
        config = yaml.load(f, Loader=yaml.FullLoader)

    config['is_adult'] = is_adult

    # Get paths for all cohorts
    cities_countries = config['cities_countries']

    # Only filter out certain cities for adult data
    if is_adult:
        cities_countries = [city_country for city_country in cities_countries if city_country != 'newcastle' and city_country != 'turkey' and city_country != 'Vessen']

    paths = [f"{root_path}{city_country}/aipal/predict.csv" for city_country in cities_countries]

    # Load and concatenate data
    df = pd.DataFrame()
    for path in paths:
        try:
            df_small = pd.read_csv(path)
            df_small['city_country'] = path.split('/')[-3]
            df = pd.concat([df, df_small])
        except Exception as e:
            print(f"Error loading {path}: {e}")

    # Filter by age based on is_adult parameter
    if is_adult:
        print("Filtering adults")
        df = df[df['age'] > 18]
    else:
        print("Filtering children")
        df = df[df['age'] <= 18]

    # remove samples with more than 20 % missing values
    df_input_features = df[config['feature_columns']]
    missing_samples = df_input_features.isna().sum(axis=1) > 0.2 * len(config['feature_columns'])
    df = df[~missing_samples]


    df['city_country'] = df['city_country'].str.replace('_', ' ')
    df['city_country'] = df['city_country'].str.capitalize()

    # map M and F to Male and Female
    df['sex'] = df['sex'].replace({'M': 'Male', 'F': 'Female'})
    df['sex'] = df['sex'].replace('I', 'Male')

    # Drop unnecessary columns
    df.drop(columns=['ELN', 'Diagnosis', 'additional.diagnosis.details..lineage.etc', 'lineage.details'],
            inplace=True, errors='ignore')

    # Clean class values
    df['class'] = df['class'].str.strip()

    # Filter cohorts by size if requested
    if filter_by_size:
        df = df.groupby('city_country').filter(lambda x: len(x) > 30)

    # Get feature columns from config
    features = config['feature_columns']

    return df, config, features

def save_to_excel(df_table, output_path):
    """Save the analysis table to a formatted Excel file with features as rows and leukemia types as columns."""
    # Create a Pandas Excel writer using openpyxl
    writer = pd.ExcelWriter(output_path, engine='openpyxl')

    # Convert the dataframe to an Excel object
    df_table.to_excel(writer, sheet_name='Analysis', index=False)

    # Get the worksheet
    worksheet = writer.sheets['Analysis']

    # Define styles
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
    feature_font = Font(bold=True)
    feature_fill = PatternFill(start_color="DBE5F1", end_color="DBE5F1", fill_type="solid")
    reference_font = Font(italic=True, color="666666")
    reference_fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
    combined_font = Font(bold=True)
    type_font = Font(italic=True)
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    # Format headers with a bold blue style and ensure proper alignment
    for col_num, column_title in enumerate(df_table.columns, 1):
        cell = worksheet.cell(row=1, column=col_num)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        cell.border = border

    # Format data cells and adjust column widths
    column_widths = {}

    # First pass to calculate optimal column widths
    for col_idx, column in enumerate(df_table.columns, 1):
        column_letter = get_column_letter(col_idx)
        # For disease type columns (may include n values), set a fixed width
        if any(class_type in column for class_type in ['ALL', 'AML', 'APL']):
            column_widths[column_letter] = 20
        elif column == 'Variable':
            column_widths[column_letter] = 18
        elif column == 'Type':
            column_widths[column_letter] = 15
        else:
            # For other columns, calculate based on content
            max_length = max(
                len(str(column)),
                df_table[column].astype(str).map(len).max()
            )
            column_widths[column_letter] = min(max_length + 2, 25)

    # Apply calculated column widths
    for col_letter, width in column_widths.items():
        worksheet.column_dimensions[col_letter].width = width

    # Format rows
    prev_variable = None
    for row_idx in range(2, worksheet.max_row + 1):
        variable = worksheet.cell(row=row_idx, column=1).value
        type_cell = worksheet.cell(row=row_idx, column=2)

        # Style format for feature rows (with variable name)
        if variable and variable.strip() and variable != prev_variable:
            # This is a feature row - bold it and color the background
            for col_idx in range(1, worksheet.max_column + 1):
                cell = worksheet.cell(row=row_idx, column=col_idx)
                cell.font = feature_font
                cell.fill = feature_fill
            prev_variable = variable

        # Style reference vs combined values differently
        if type_cell.value:
            type_value = type_cell.value.strip()
            if type_value.startswith("Reference"):
                # Format reference rows with gray italic text
                for col_idx in range(1, worksheet.max_column + 1):
                    cell = worksheet.cell(row=row_idx, column=col_idx)
                    cell.font = reference_font
                    cell.fill = reference_fill
            elif type_value == "Mean (sd)" or type_value == "Median [IQR]":
                # Format combined data rows with bold text
                for col_idx in range(1, worksheet.max_column + 1):
                    cell = worksheet.cell(row=row_idx, column=col_idx)
                    if col_idx >= 3:  # Only the actual data cells, not the labels
                        cell.font = combined_font
                    else:
                        cell.font = type_font
            else:
                # Make the type cell italic
                type_cell.font = type_font

        # Center-align all cells and add borders
        for col_idx in range(1, worksheet.max_column + 1):
            cell = worksheet.cell(row=row_idx, column=col_idx)
            # Align differently based on column
            if col_idx == 1:  # Variable column
                cell.alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
            elif col_idx == 2:  # Type column
                cell.alignment = Alignment(horizontal='right', vertical='center', wrap_text=True)
            else:  # Data columns (ALL, AML, APL)
                cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            cell.border = border

    # Add thick horizontal lines between feature groups
    prev_variable = None
    for row_idx in range(2, worksheet.max_row + 1):
        variable = worksheet.cell(row=row_idx, column=1).value

        # Add a thicker border above feature rows
        if variable and variable.strip() and variable != prev_variable and row_idx > 2:
            for col_idx in range(1, worksheet.max_column + 1):
                cell = worksheet.cell(row=row_idx, column=col_idx)
                thick_border = Border(
                    left=Side(style='thin'),
                    right=Side(style='thin'),
                    top=Side(style='thick'),
                    bottom=Side(style='thin')
                )
                cell.border = thick_border

        prev_variable = variable if variable else prev_variable

    # Freeze the header row and first two columns
    worksheet.freeze_panes = 'C2'

    # Save the workbook
    writer.close()

df, config, features = load_data()

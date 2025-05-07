# %%
import pandas as pd
import yaml
import ast
import numpy as np
import os
import re
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side, Color
from openpyxl.utils import get_column_letter
from openpyxl.styles.differential import DifferentialStyle
from openpyxl.formatting.rule import Rule
from tabulate import tabulate
import openpyxl

# Get the script's directory
script_dir = os.path.dirname(os.path.abspath(__file__))
config_path = os.path.join(script_dir, '../aipal_validation/config/config_outlier.yaml')

# Load configuration and set paths
root_path = '/local/work/merengelke/aipal/'
config = yaml.load(open(config_path, 'r'), Loader=yaml.FullLoader)

cities_countries = config['cities_countries']
paths = [f"{root_path}{city_country}/aipal/results.csv" for city_country in cities_countries]
results_paths = [f"{root_path}{city_country}/aipal/predict.csv" for city_country in cities_countries]

# Load and combine data with proper handling of empty cells
df = pd.DataFrame()
for path in paths:
    df_small = pd.read_csv(path)
    df_small['city_country'] = path.split('/')[-3]
    # Fill empty cells with NaN
    df_small = df_small.replace('', np.nan)
    df = pd.concat([df, df_small])

df_results = pd.DataFrame()
for path in results_paths:
    df_small = pd.read_csv(path)
    df_small['city_country'] = path.split('/')[-3]
    df_results = pd.concat([df_results, df_small])

# Define cities to exclude for adult analysis
excluded_cities = ['Vessen', 'newcastle', 'turkey']

# Filter based on age group
if config['is_adult']:
    df = df.loc[:, ~df.columns.str.contains('kids')]
    # Remove specific cities for adult analysis
    df = df[~df['city_country'].isin(excluded_cities)]
else:
    df = df.loc[:, ~df.columns.str.contains('adults')]
    # No city exclusions for kids analysis

df.rename(columns={'Unnamed: 0': 'class'}, inplace=True)

def safe_extract_metric(row, metric_name):
    """Extract metric values from the dictionary-like string in the CSV"""
    try:
        # Convert row to string if it's not already
        row_str = str(row)

        # Use regex to extract the values for the specific metric
        # The pattern matches: 'MetricName': [value1, value2, ...]
        pattern = f"'{metric_name}':\s*\[(.*?)\]"
        match = re.search(pattern, row_str)

        if match:
            values_str = match.group(1)
            # Split by comma and convert to floats, excluding nan values
            values = [float(v.strip()) for v in values_str.split(',')
                     if v.strip().lower() != 'nan']

            if not values:
                return None, None

            mean_value = np.mean(values)
            ci = None
            if len(values) > 1:
                ci = np.std(values) * 1.96 / np.sqrt(len(values))
            return mean_value, ci
        return None, None
    except Exception as e:
        print(f"Error extracting {metric_name}: {e}")
        print(f"Row content: {row}")
        return None, None

# Metrics and cutoff types
metrics = ['AUC', 'Accuracy', 'Precision', 'Recall', 'F1 Score']

# Make cutoff types dynamic based on is_adult setting
if config['is_adult']:
    cutoff_types = ['no cutoff - adults', 'overall cutoff - adults', 'confident cutoff - adults']
    age_suffix = 'adults'
else:
    cutoff_types = ['no cutoff - kids', 'overall cutoff - kids', 'confident cutoff - kids']
    age_suffix = 'kids'

cutoff_names = ['No Cutoff', 'Overall Cutoff', 'Confident Cutoff']
classes = ['AML', 'APL', 'ALL']

# Create results directories
tables_dir = os.path.join(script_dir, 'tables')
excel_dir = os.path.join(script_dir, 'excel_tables')
os.makedirs(excel_dir, exist_ok=True)
os.makedirs(tables_dir, exist_ok=True)  # For markdown files

def style_worksheet(ws):
    """Apply professional styling to the Excel worksheet"""
    # Define custom colors
    header_color = "366092"  # Dark blue
    alt_row_color = "E6EFF7"  # Light blue
    border_color = "BFBFBF"  # Light gray
    metric_font_color = "000000"  # Black

    # Define styles
    header_fill = PatternFill(start_color=header_color, end_color=header_color, fill_type="solid")
    alt_row_fill = PatternFill(start_color=alt_row_color, end_color=alt_row_color, fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=12)
    metric_font = Font(bold=True, color=metric_font_color, size=11)
    data_font = Font(size=11)

    # Define borders
    thin_border = Border(
        left=Side(style='thin', color=border_color),
        right=Side(style='thin', color=border_color),
        top=Side(style='thin', color=border_color),
        bottom=Side(style='thin', color=border_color)
    )

    bottom_border = Border(
        left=Side(style='thin', color=border_color),
        right=Side(style='thin', color=border_color),
        top=Side(style='thin', color=border_color),
        bottom=Side(style='medium', color=border_color)
    )

    # Style headers
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = thin_border

    # Style data rows with alternating colors
    row_count = len(list(ws.rows))
    for idx, row in enumerate(ws.iter_rows(min_row=2, max_row=row_count)):
        # First column (metrics) gets special styling
        row[0].font = metric_font
        row[0].alignment = Alignment(horizontal='left', vertical='center')

        # Style all cells in the row
        for cell in row:
            cell.border = thin_border
            if idx % 2 == 0:  # Even rows get background color
                cell.fill = alt_row_fill
            else:
                cell.fill = PatternFill(fill_type=None)

            # Style number cells
            if cell.column > 1:  # Data cells (not metric names)
                cell.font = data_font
                cell.alignment = Alignment(horizontal='center', vertical='center')

                # Format the values for better readability
                if cell.value != "N/A":
                    try:
                        # If the cell contains a value and CI like "0.84 (0.84-0.85)"
                        if '(' in str(cell.value):
                            parts = cell.value.split('(')
                            mean = float(parts[0].strip())
                            cell.value = f"{mean:.2f} ({parts[1]}"
                    except:
                        pass

    # Add a thicker bottom border to the last row
    for cell in ws[row_count]:
        cell.border = bottom_border

    # Adjust column widths
    ws.column_dimensions[get_column_letter(1)].width = 15  # Metric column
    for col_idx in range(2, len(classes) + 2):
        ws.column_dimensions[get_column_letter(col_idx)].width = 16  # Data columns

    # Freeze the header row
    ws.freeze_panes = 'A2'

    # Add auto-filter to header row
    ws.auto_filter.ref = f"A1:{get_column_letter(len(classes) + 1)}1"

def generate_markdown_table(city_country, cutoff_name, metrics, classes, results, cutoff):
    """Generate a visually improved markdown table"""
    # Title with nice formatting
    markdown_content = [f"# Performance Metrics for {city_country.capitalize()} - {cutoff_name.title()}\n\n"]

    # Brief description
    description = (
        f"This table presents the model performance metrics for {city_country.capitalize()} "
        f"with {cutoff_name.lower()} threshold settings. Each metric is presented as Mean (95% CI).\n\n"
    )
    markdown_content.append(description)

    # Table header with better alignment
    markdown_content.append("| Metric    | " + " | ".join([f"{cls}" for cls in classes]) + " |\n")
    markdown_content.append("|:----------|:" + "|:".join(["------:" for _ in classes]) + "|\n")

    # Table content with consistent formatting
    for metric in metrics:
        row = [f"**{metric}**"]  # Bold metric names
        for class_name in classes:
            if (cutoff in results and
                class_name in results[cutoff] and
                metric in results[cutoff][class_name]):
                value = results[cutoff][class_name][metric]
                # Format with consistent decimal places
                formatted_value = f"{value['mean']:.2f} ({value['ci'][0]:.2f}-{value['ci'][1]:.2f})"
                row.append(formatted_value)
            else:
                row.append("N/A")
        markdown_content.append("| " + " | ".join(row) + " |\n")

    # Add a footer note
    markdown_content.append("\n*Note: N/A indicates that data is not available for this metric.*\n")

    return "".join(markdown_content)

def process_results(df, city_country):
    # Use the dynamic cutoff types based on is_adult
    if config['is_adult']:
        cutoff_types = ['no cutoff - adults', 'overall cutoff - adults', 'confident cutoff - adults']
        age_suffix = 'adults'
    else:
        cutoff_types = ['no cutoff - kids', 'overall cutoff - kids', 'confident cutoff - kids']
        age_suffix = 'kids'

    classes = ['AML', 'APL', 'ALL']
    metrics = ['AUC', 'Accuracy', 'Precision', 'Recall', 'F1 Score']

    results = {}
    for cutoff in cutoff_types:
        results[cutoff] = {}
        for class_name in classes:
            results[cutoff][class_name] = {}
            try:
                # Get all rows for this class
                class_data = df[df['class'] == class_name]
                if class_data.empty:
                    continue

                # Find all non-NaN values for this cutoff
                valid_rows = class_data[~pd.isna(class_data[cutoff])]
                if valid_rows.empty:
                    continue

                # Process the first valid row
                row = valid_rows.iloc[0][cutoff]

                for metric in metrics:
                    value, ci = safe_extract_metric(row, metric)
                    if value is not None:
                        results[cutoff][class_name][metric] = {
                            'mean': value,
                            'ci': [value - ci if ci else value, value + ci if ci else value]
                        }
            except Exception as e:
                print(f"Error processing {class_name} for {cutoff}: {e}")
                continue

    # Generate Excel tables
    wb = Workbook()
    wb.remove(wb.active)  # Remove default sheet

    # Generate Markdown tables
    os.makedirs(tables_dir, exist_ok=True)

    for cutoff in cutoff_types:
        cutoff_name = cutoff.replace(f' - {age_suffix}', '')
        sheet_name = f"{city_country}_{cutoff_name}"
        ws = wb.create_sheet(sheet_name)

        # Add a title to the Excel sheet
        ws.title = sheet_name

        # Write headers
        headers = ['Metric'] + classes
        ws.append(headers)

        # Write data
        for metric in metrics:
            row = [metric]
            for class_name in classes:
                if (cutoff in results and
                    class_name in results[cutoff] and
                    metric in results[cutoff][class_name]):
                    value = results[cutoff][class_name][metric]
                    formatted_value = f"{value['mean']:.2f} ({value['ci'][0]:.2f}-{value['ci'][1]:.2f})"
                    row.append(formatted_value)
                else:
                    row.append("N/A")
            ws.append(row)

        # Style the Excel worksheet
        style_worksheet(ws)

        # Generate and save the markdown table
        cutoff_underscore = cutoff_name.replace(' ', '_')
        markdown_content = generate_markdown_table(city_country, cutoff_name, metrics, classes, results, cutoff)
        markdown_filename = os.path.join(tables_dir, f"{city_country}_{cutoff_underscore}.md")

        with open(markdown_filename, 'w') as f:
            f.write(markdown_content)

        print(f"Created markdown file: {markdown_filename}")

    return wb

def apply_summary_styling(summary_sheet, summary_rows):
    """Apply styling to the summary sheet"""
    # Headers
    header_color = "366092"  # Dark blue
    alt_row_color = "E6EFF7"  # Light blue
    border_color = "BFBFBF"  # Light gray

    header_fill = PatternFill(start_color=header_color, end_color=header_color, fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=12)

    for cell in summary_sheet[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center', vertical='center')

    # Adjust column widths
    summary_sheet.column_dimensions['A'].width = 20
    summary_sheet.column_dimensions['B'].width = 15
    summary_sheet.column_dimensions['C'].width = 30

    # Add alternating row colors
    alt_row_fill = PatternFill(start_color=alt_row_color, end_color=alt_row_color, fill_type="solid")
    for idx in range(2, summary_rows + 1):
        if idx % 2 == 0:  # Even rows
            for cell in summary_sheet[idx]:
                cell.fill = alt_row_fill

    # Add borders
    thin_border = Border(
        left=Side(style='thin', color=border_color),
        right=Side(style='thin', color=border_color),
        top=Side(style='thin', color=border_color),
        bottom=Side(style='thin', color=border_color)
    )
    for row in summary_sheet.iter_rows(min_row=1, max_row=summary_rows):
        for cell in row:
            cell.border = thin_border

    # Freeze panes and autofilter
    summary_sheet.freeze_panes = 'A2'
    summary_sheet.auto_filter.ref = f"A1:C{summary_rows}"

def main():
    # Create output directories
    os.makedirs(excel_dir, exist_ok=True)
    os.makedirs(tables_dir, exist_ok=True)

    # Create a summary workbook
    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    # Create a summary sheet
    summary_sheet = wb.create_sheet("Summary")
    summary_sheet.append(["City", "Cutoff Type", "Data Available"])
    summary_rows = 1  # Keep track of rows for styling

    # Filter cities based on adult/child analysis
    cities_to_process = cities_countries.copy()
    if config['is_adult']:
        # Remove excluded cities for adult analysis
        cities_to_process = [city for city in cities_to_process if city not in excluded_cities]

    # Track sheets to add to the main workbook
    sheets_to_add = []

    # Process all cities from the filtered list
    for city_country in cities_to_process:
        print(f"\nProcessing {city_country}:")
        results_path = os.path.join('/local/work/merengelke/aipal', city_country, 'aipal', 'results.csv')

        try:
            # Skip cities that don't have a results.csv file
            if not os.path.exists(results_path):
                print(f"No results.csv file found for {city_country}")
                summary_sheet.append([city_country.capitalize(), "All", "No data available"])
                summary_rows += 1
                continue

            df_city = pd.read_csv(results_path)
            df_city.rename(columns={'Unnamed: 0': 'class'}, inplace=True)

            # Apply same filtering logic as in the beginning of the script
            if config['is_adult']:
                df_city = df_city.loc[:, ~df_city.columns.str.contains('kids')]
            else:
                df_city = df_city.loc[:, ~df_city.columns.str.contains('adults')]

            # Check if there are any data columns left after filtering
            data_columns = [col for col in df_city.columns if col not in ['class', 'city_country']]
            if not data_columns:
                print(f"No {'adult' if config['is_adult'] else 'kids'} data for {city_country}, skipping")
                summary_sheet.append([city_country.capitalize(), "All", f"No {'adult' if config['is_adult'] else 'kids'} data available"])
                summary_rows += 1
                continue

            # Process the results
            city_wb = process_results(df_city, city_country)

            # Update the summary
            for cutoff_name in ["No Cutoff", "Overall Cutoff", "Confident Cutoff"]:
                summary_sheet.append([city_country.capitalize(), cutoff_name, "Yes"])
                summary_rows += 1

            # Add sheets info to the list
            for sheet_name in city_wb.sheetnames:
                source_ws = city_wb[sheet_name]
                sheets_to_add.append({
                    'name': sheet_name,
                    'data': [
                        [cell.value for cell in row]
                        for row in source_ws.rows
                    ]
                })

        except Exception as e:
            print(f"Error processing {city_country}: {str(e)}")
            summary_sheet.append([city_country.capitalize(), "All", f"Error: {str(e)[:50]}"])
            summary_rows += 1
            continue

    # Add all sheets to the main workbook
    for sheet_info in sheets_to_add:
        ws = wb.create_sheet(sheet_info['name'])

        # Add the data
        for row_data in sheet_info['data']:
            ws.append(row_data)

        # Apply styling to this sheet
        style_worksheet(ws)

    # Apply styling to the summary sheet
    apply_summary_styling(summary_sheet, summary_rows)

    # Make the summary sheet the first sheet
    wb.move_sheet("Summary", 0)

    # Save Excel file
    if config['is_adult']:
        excel_path = os.path.join(excel_dir, f'cv_results_adult.xlsx')
    else:
        excel_path = os.path.join(excel_dir, f'cv_results_child.xlsx')
    wb.save(excel_path)
    print(f"\nAll tables have been saved to {excel_path} and as individual .md files in the {tables_dir} directory")

if __name__ == '__main__':
    main()
